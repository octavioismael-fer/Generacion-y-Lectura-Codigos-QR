# Importamos Librerias 
import cv2
from pyzbar.pyzbar import decode
import numpy as np
from datetime import datetime
import openpyxl as xl

# Creo VieoCaputura
cap = cv2.VideoCapture(0)

# Varibales (Son las hojas de excel para los dintitos horario de los dias)
morning = []
afternoon = []
night = []

# Horario
def infhora():
    # Informacion
    inf = datetime.now()
    # Extraemos Fecha
    fecha = inf.strftime('%Y:%m:%d')
    # Extraemos Hora
    hora = inf.strftime('%H:%M:%S')

    return hora, fecha

while True:
    # Leemos los frames de la camara
    ret, frame = cap.read()

    # Interfaz principal (mostrara un cuadrado verte que indica el lugar donde se debe colocar el qr para poder leerlo y un breve texto)
    # Texto
    cv2.putText(frame, 'Coloque el Codigo QR', (160,80), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)
    # Ubicar el regtangulo
    cv2.rectangle(frame, (170, 100), (470, 400), (0, 255, 0), 2)

    # Extraemos hora y fecha
    hora, fecha =   infhora()
    diasem = datetime.today().weekday()

    print(diasem)
    # AÑO - MES - DIA
    a, me, d = fecha[0:4], fecha[5:7], fecha[8:10]
    # HORA - MINUTO - SEGUNDO
    h, m, s = int(hora[0:2]), int(hora[3:5]), int(hora[6:8])

    # Creamos Archivos excel
    nomar = str(a) + '-' + str(me) + '-' + str(d)
    texth = str(h) + '-' + str(m) + '-' + str(s)
    print(nomar)
    print(texth)
    # Archivo Excel
    wb = xl.Workbook()

    # Leemos los codigos QR
    for codes in decode(frame):

        # INFORMACION
        # Decodificacion
        info = codes.data.decode('utf-8')

        # Tipo de persona LETRA
        tipo = info[0:2]
        tipo = int(tipo)
        letr = chr(tipo)

        # Numero
        num = info[2:]

        # Extraemos las coordenadas
        pts = np.array([codes.polygon], np.int32)
        xi, yi = codes.rect.left, codes.rect.top

        # Redimensionamos
        pts = pts.reshape((-1, 1, 2))

        # ID Completo
        codigo = letr + num

        # DIAS DE LA SEMANA
        # SEMANAS
        if 4 >= diasem >= 0:

            # Divido las horas del dia
            # Mañana (morning)
            if 12 >= h >= 5:
                cv2.polylines(frame, [pts], True, (255, 255, 0), 5)
                # Guardamos el ID
                if codigo not in morning:
                    # Agregamos ID
                    pos = len(morning)
                    morning.append(codigo)

                    # Guardamos DataBase
                    hojam = wb.create_sheet("Mañana")
                    datos = hojam.append(morning)
                    wb.save(nomar + '.xlsx')

                    # Dibujamos el contorno del QR
                    cv2.putText(frame, letr + '0' + str(num), (xi - 15, yi - 15), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 55, 0), 2)
                
                # AVISAMOS
                elif codigo in morning:
                    cv2.putText(frame, 'EL ID ' + str(codigo), (xi - 65, yi - 45), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 55, 0), 2)
                    cv2.putText(frame, 'Fue Registrado', (xi - 65, yi - 15), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 55, 0), 2)
            
            # Tarde (afternoon)
            if 20 >= h >= 13:
                cv2.polylines(frame, [pts], True, (255, 255, 0), 5)
                # Guardamos el ID
                if codigo not in afternoon:
                    # Agregamos ID
                    afternoon.append(codigo)
                    #Dibujamos el contorno del QR
                    cv2.putText(frame, letr + '0' + str(num), (xi - 15, yi - 15), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 55, 0), 2)

                    # Guardamos DataBase
                    hojam = wb.create_sheet("Tarde")
                    datos = hojam.append(morning)
                    wb.save(nomar + '.xlsx')
                
                # AVISAMOS
                elif codigo in morning:
                    cv2.putText(frame, 'EL ID ' + str(codigo), (xi - 65, yi - 45), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 55, 0), 2)
                    cv2.putText(frame, 'Fue Registrado', (xi - 65, yi - 15), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 55, 0), 2)
            
            # Noche (nigth)
            if 23 >= h >= 21:
                cv2.polylines(frame, [pts], True, (255, 255, 0), 5)
                # Guardamos el ID
                if codigo not in night:
                    # Agregamos ID
                    night.append(codigo)
                    #Dibujamos el contorno del QR
                    cv2.putText(frame, letr + '0' + str(num), (xi - 15, yi - 15), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 55, 0), 2)

                    # Guardamos DataBase
                    hojam = wb.create_sheet("Tarde")
                    datos = hojam.append(morning)
                    wb.save(nomar + '.xlsx')
                
                # AVISAMOS
                elif codigo in morning:
                    cv2.putText(frame, 'EL ID ' + str(codigo), (xi - 65, yi - 45), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 55, 0), 2)
                    cv2.putText(frame, 'Fue Registrado', (xi - 65, yi - 15), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 55, 0), 2)

    # Mostramos FPS
    cv2.imshow( " LECTOR DE QR ", frame)
    # Leemos el teclado
    t = cv2.waitKey(5)
    if t == 27:
        break

cv2.destroyAllWindows()
cap.release()